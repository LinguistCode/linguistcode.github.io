#!/usr/bin/env python3
"""
Analyse la valence (positif/négatif) et l'émotion de collocations anglaises,
par comparaison à des dictionnaires lexicaux (approche lexicon-based).

Colonne 1 (A) : "collocation"  -> lue, JAMAIS modifiée
Colonne 8 (H) : Valence        -> écrite/écrasée
Colonne 9 (I) : Émotion        -> écrite/écrasée
Toutes les autres colonnes sont laissées strictement intactes.

Méthode (en cascade, pour maximiser la couverture ET la fiabilité) :
1. VADER (vaderSentiment) : lexique de sentiment ~7500 mots, gère les
   négations/intensificateurs ("not good", "very bad"). Essayé en premier
   car le plus fin.
2. SentiWordNet (via NLTK/WordNet) : lexique beaucoup plus large
   (~117 000 mots), utilisé UNIQUEMENT si aucun mot de l'expression n'est
   dans VADER. Chaque mot est lemmatisé et étiqueté grammaticalement
   (POS-tagging) pour choisir le bon sens WordNet.
3. "Non trouvé" seulement si NI VADER NI SentiWordNet ne connaissent le mot.
La source utilisée est indiquée dans la cellule, pour auditer la fiabilité.

Émotion (secondaire) : NRCLex (NRC Word-Emotion Association Lexicon).
Best-effort : si rien n'est trouvé ou si le module échoue, le script
continue sans bloquer l'analyse de valence (qui reste prioritaire).

Installation :
    python -m pip install openpyxl vaderSentiment nrclex nltk
(les données NLTK nécessaires - wordnet, sentiwordnet, tagger - sont
téléchargées automatiquement au premier lancement : connexion internet
requise la première fois)

Usage :
    python analyse_valence_collocations.py entree.xlsx sortie.xlsx
"""

import sys
import re
import openpyxl
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

# --- Téléchargement (silencieux, best-effort) des ressources NLTK ----------
import nltk
for _res in ("punkt", "punkt_tab", "wordnet", "omw-1.4", "sentiwordnet",
             "averaged_perceptron_tagger", "averaged_perceptron_tagger_eng"):
    try:
        nltk.download(_res, quiet=True)
    except Exception:
        pass

from nltk import pos_tag
from nltk.corpus import wordnet as wn
from nltk.corpus import sentiwordnet as swn
from nltk.stem import WordNetLemmatizer

LEMMATIZER = WordNetLemmatizer()

# --- Émotion (secondaire) : NRCLex, dépendance optionnelle ------------------
try:
    from nrclex import NRCLex
    NRCLEX_OK = True
except ImportError:
    NRCLEX_OK = False

# --- Paramètres --------------------------------------------------------
COLONNE_ENTREE = 1      # A : collocation
COLONNE_VALENCE = 8     # H
COLONNE_EMOTION = 9     # I

# Seuils recommandés par les auteurs de VADER pour le score "compound"
# (conservés aussi pour SentiWordNet, échelle comparable -1..+1)
SEUIL_POSITIF = 0.05
SEUIL_NEGATIF = -0.05

EMOTIONS_FR = {
    "fear": "Peur", "anger": "Colère", "anticip": "Anticipation",
    "anticipation": "Anticipation", "trust": "Confiance", "surprise": "Surprise",
    "sadness": "Tristesse", "disgust": "Dégoût", "joy": "Joie",
}


def extraire_mots(texte):
    """Tokenisation simple : suites de lettres/apostrophes, en minuscules."""
    return re.findall(r"[a-zA-Z']+", str(texte).lower())


def treebank_vers_wordnet(tag):
    """Convertit un tag POS (NLTK/Penn Treebank) vers la constante WordNet."""
    if tag.startswith("J"):
        return wn.ADJ
    if tag.startswith("V"):
        return wn.VERB
    if tag.startswith("N"):
        return wn.NOUN
    if tag.startswith("R"):
        return wn.ADV
    return None


def score_sentiwordnet(texte):
    """
    Score de valence via SentiWordNet, moyenné sur les mots reconnus.
    Chaque mot est lemmatisé puis on prend le sens WordNet le plus fréquent
    (première entrée), dont le score = pos_score - neg_score (échelle -1..+1,
    l'objectivité/neutralité venant réduire ce score vers 0).
    Retourne None si aucun mot n'est reconnu.
    """
    mots = extraire_mots(texte)
    if not mots:
        return None
    try:
        tags = pos_tag(mots)
    except Exception:
        tags = [(m, "NN") for m in mots]  # repli : traiter tout comme nom

    scores = []
    for mot, tag in tags:
        wn_pos = treebank_vers_wordnet(tag)
        if wn_pos is None:
            continue
        try:
            lemme = LEMMATIZER.lemmatize(mot, pos=wn_pos)
            synsets = list(swn.senti_synsets(lemme, wn_pos))
        except Exception:
            synsets = []
        if not synsets:
            continue
        sens_principal = synsets[0]  # sens le plus fréquent dans WordNet
        scores.append(sens_principal.pos_score() - sens_principal.neg_score())

    if not scores:
        return None
    return sum(scores) / len(scores)


def analyser_valence(texte, analyzer):
    """
    Retourne (label, score, source).
    label : "Positif" / "Négatif" / "Neutre" / "Non trouvé".
    source : "VADER" ou "SentiWordNet" (ou None si "Non trouvé").
    """
    mots = extraire_mots(texte)
    mots_connus_vader = [m for m in mots if m in analyzer.lexicon]

    if mots_connus_vader:
        compound = analyzer.polarity_scores(str(texte))["compound"]
        source = "VADER"
    else:
        compound = score_sentiwordnet(texte)
        source = "SentiWordNet"

    if compound is None:
        return "Non trouvé", None, None

    if compound >= SEUIL_POSITIF:
        label = "Positif"
    elif compound <= SEUIL_NEGATIF:
        label = "Négatif"
    else:
        label = "Neutre"

    return label, compound, source


def analyser_emotion(texte):
    """Émotion dominante selon NRCLex (analyse secondaire, best-effort)."""
    if not NRCLEX_OK:
        return "N/A (nrclex non installé)"
    try:
        obj = NRCLex(str(texte))
        freqs = {k: v for k, v in obj.affect_frequencies.items()
                 if k not in ("positive", "negative") and v > 0}

        if not freqs:
            # Repli : tenter avec les lemmes (forme de base), qui matchent
            # parfois mieux le lexique NRC que la forme fléchie d'origine.
            mots = extraire_mots(texte)
            try:
                tags = pos_tag(mots)
            except Exception:
                tags = [(m, "NN") for m in mots]
            lemmes = []
            for mot, tag in tags:
                wn_pos = treebank_vers_wordnet(tag) or wn.NOUN
                try:
                    lemmes.append(LEMMATIZER.lemmatize(mot, pos=wn_pos))
                except Exception:
                    lemmes.append(mot)
            obj2 = NRCLex(" ".join(lemmes))
            freqs = {k: v for k, v in obj2.affect_frequencies.items()
                     if k not in ("positive", "negative") and v > 0}

        if not freqs:
            return "Aucune"
        top = max(freqs, key=freqs.get)
        return EMOTIONS_FR.get(top, top.capitalize())
    except Exception:
        return "Indéterminée"


def main(fichier_entree, fichier_sortie):
    wb = openpyxl.load_workbook(fichier_entree)
    ws = wb.active
    analyzer = SentimentIntensityAnalyzer()

    entete = ws.cell(row=1, column=COLONNE_ENTREE).value
    if not entete or "collocation" not in str(entete).lower():
        print(f"Avertissement : la colonne 1 s'intitule '{entete}', pas "
              f"'collocation'. Poursuite sur la colonne 1 telle quelle.")

    ws.cell(row=1, column=COLONNE_VALENCE, value="Valence")
    ws.cell(row=1, column=COLONNE_EMOTION, value="Émotion")

    total = positifs = negatifs = neutres = non_trouves = 0
    par_source = {"VADER": 0, "SentiWordNet": 0}

    for row in range(2, ws.max_row + 1):
        collocation = ws.cell(row=row, column=COLONNE_ENTREE).value
        if collocation is None or str(collocation).strip() == "":
            continue

        total += 1
        label, score, source = analyser_valence(collocation, analyzer)
        emotion = analyser_emotion(collocation)

        if label == "Non trouvé":
            non_trouves += 1
            valeur_cellule = "Non trouvé"
        else:
            valeur_cellule = f"{label} ({score:+.3f}, {source})"
            par_source[source] += 1
            if label == "Positif":
                positifs += 1
            elif label == "Négatif":
                negatifs += 1
            else:
                neutres += 1

        ws.cell(row=row, column=COLONNE_VALENCE, value=valeur_cellule)
        ws.cell(row=row, column=COLONNE_EMOTION, value=emotion)

    wb.save(fichier_sortie)

    print(f"Terminé : {fichier_sortie}")
    print(f"  Lignes analysées : {total}")
    print(f"  Positif / Négatif / Neutre : {positifs} / {negatifs} / {neutres}")
    if total:
        print(f"  Non trouvées (ni VADER ni SentiWordNet) : {non_trouves} "
              f"({non_trouves / total * 100:.1f} %)")
        print(f"  Trouvées via VADER : {par_source['VADER']} | "
              f"via SentiWordNet : {par_source['SentiWordNet']}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage : python analyse_valence_collocations.py entree.xlsx sortie.xlsx")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])