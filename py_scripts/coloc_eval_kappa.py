#!/usr/bin/env python3
"""
Compare l'annotation automatique (colonne "Valence", ex. produite par
analyse_valence_collocations.py) à une annotation manuelle de contrôle
(colonne "Verification"), pour évaluer la fiabilité du script.

Calcule :
- le taux d'accord simple (% de lignes où les deux annotations concordent)
- le kappa de Cohen (accord corrigé du hasard, mesure standard en
  linguistique de corpus / annotation de contenu)
- une matrice de confusion (pour voir sur quelles paires de labels
  portent les désaccords : ex. Neutre confondu avec Positif, etc.)

Ne modifie AUCUNE colonne du fichier : lecture seule, résultats affichés
dans la console (et, en option, exportés dans un fichier texte).

Le script repère les colonnes par leur EN-TÊTE ("Valence" et
"Verification"), peu importe leur position exacte dans le tableau.
La colonne "Valence" peut contenir du texte du type
"Positif (+0.620, VADER)" : seul le premier mot (le label) est comparé.

Installation :
    python -m pip install openpyxl

Usage :
    python comparer_annotations.py fichier.xlsx
"""

import sys
import re
import openpyxl
from collections import Counter


def extraire_label(valeur):
    """
    Isole le label (Positif/Négatif/Neutre/Non trouvé) à partir d'une
    cellule qui peut être soit déjà un label simple, soit une chaîne du
    type "Positif (+0.620, VADER)".
    """
    if valeur is None:
        return None
    texte = str(valeur).strip()
    if not texte:
        return None
    # Le label est le ou les premiers mots avant une parenthèse éventuelle
    match = re.match(r"^([A-Za-zÀ-ÿ]+(?:\s[A-Za-zÀ-ÿ]+)?)", texte)
    if not match:
        return texte
    return match.group(1).strip()


def trouver_colonne(ws, nom_recherche):
    """Retourne le numéro de colonne dont l'en-tête (ligne 1) correspond
    au nom recherché (insensible à la casse/accents approximative)."""
    for col in range(1, ws.max_column + 1):
        entete = ws.cell(row=1, column=col).value
        if entete and nom_recherche.lower() in str(entete).lower():
            return col
    return None


def kappa_cohen(paires):
    """
    Calcule le kappa de Cohen à partir d'une liste de tuples
    (label_auto, label_manuel).
    kappa = (accord_observé - accord_attendu) / (1 - accord_attendu)
    """
    n = len(paires)
    if n == 0:
        return None

    accord_observe = sum(1 for a, m in paires if a == m) / n

    labels = sorted(set([a for a, m in paires] + [m for a, m in paires]))
    compte_auto = Counter(a for a, m in paires)
    compte_manuel = Counter(m for a, m in paires)

    accord_attendu = sum(
        (compte_auto[l] / n) * (compte_manuel[l] / n) for l in labels
    )

    if accord_attendu == 1:
        return 1.0  # cas limite : un seul label partout

    kappa = (accord_observe - accord_attendu) / (1 - accord_attendu)
    return kappa


def interpretation_kappa(k):
    """Grille de Landis & Koch (1977), largement citée en linguistique."""
    if k < 0:
        return "désaccord (pire qu'un accord aléatoire)"
    if k < 0.20:
        return "accord très faible"
    if k < 0.40:
        return "accord faible"
    if k < 0.60:
        return "accord modéré"
    if k < 0.80:
        return "accord substantiel"
    return "accord quasi parfait"


def main(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb.active

    col_valence = trouver_colonne(ws, "Valence")
    col_verif = trouver_colonne(ws, "Verification")
    # tolère aussi l'orthographe accentuée "Vérification"
    if col_verif is None:
        col_verif = trouver_colonne(ws, "Vérification")

    if col_valence is None or col_verif is None:
        print("Erreur : impossible de trouver les colonnes 'Valence' et/ou "
              "'Verification' (vérifie l'en-tête en ligne 1).")
        sys.exit(1)

    paires = []
    lignes_ignorees = 0

    for row in range(2, ws.max_row + 1):
        auto = extraire_label(ws.cell(row=row, column=col_valence).value)
        manuel = extraire_label(ws.cell(row=row, column=col_verif).value)

        if auto is None or manuel is None:
            continue  # ligne pas encore évaluée manuellement : on l'ignore
        if auto == "Non trouvé":
            # Cas particulier : le script n'a rien pu évaluer. On le compte
            # à part plutôt que de fausser l'accord/désaccord.
            lignes_ignorees += 1
            continue

        paires.append((auto, manuel))

    if not paires:
        print("Aucune paire (Valence, Verification) exploitable trouvée.")
        sys.exit(1)

    n = len(paires)
    accords = [(a, m) for a, m in paires if a == m]
    desaccords = [(a, m) for a, m in paires if a != m]
    taux_accord = len(accords) / n * 100
    kappa = kappa_cohen(paires)

    # Matrice de confusion : lignes = manuel (référence), colonnes = auto
    labels = sorted(set([a for a, m in paires] + [m for a, m in paires]))
    matrice = {l: Counter() for l in labels}
    for a, m in paires:
        matrice[m][a] += 1

    # --- Affichage -----------------------------------------------------
    print(f"Fichier : {fichier}")
    print(f"Lignes comparées : {n} "
          f"(+ {lignes_ignorees} ignorées car 'Non trouvé' par le script)")
    print()
    print(f"Taux d'accord simple : {taux_accord:.1f} %  "
          f"({len(accords)}/{n})")
    print(f"Kappa de Cohen        : {kappa:.3f}  "
          f"({interpretation_kappa(kappa)})")
    print()

    print("Matrice de confusion (lignes = annotation manuelle, "
          "colonnes = script) :")
    entete = "".ljust(14) + "".join(l.ljust(12) for l in labels)
    print(entete)
    for l_manuel in labels:
        ligne = l_manuel.ljust(14)
        for l_auto in labels:
            ligne += str(matrice[l_manuel][l_auto]).ljust(12)
        print(ligne)
    print()

    if desaccords:
        print(f"Détail des {len(desaccords)} désaccords (script -> manuel) :")
        compte_desaccords = Counter(desaccords)
        for (a, m), c in compte_desaccords.most_common():
            print(f"  {a} -> {m} : {c} cas")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage : python comparer_annotations.py fichier.xlsx")
        sys.exit(1)
    main(sys.argv[1])
